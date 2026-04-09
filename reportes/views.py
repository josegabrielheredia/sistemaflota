from datetime import date, timedelta

from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from choferes.models import Chofer, Conduce
from inventario.models import MovimientoInventario, Producto
from pagos.models import AvanceChofer, Pago
from proveedores.models import RegistroProveedor
from recursos_humanos.models import (
    Empleado,
    Licencia,
    PagoEmpleado,
    RegistroGasto,
    TipoLicencia,
    Vacacion,
)
from tracking.models import AlquilerContenedor, Contenedor, Vehiculo

from .forms import GeneradorReporteForm


def _format_currency(value):
    return f"RD$ {value:,.2f}"


def _choice_label(choices, value, default="No definido"):
    mapping = dict(choices)
    return mapping.get(value, value or default)


def _slugify_filename(value):
    normalized = "".join(char.lower() if char.isalnum() else "_" for char in value)
    return "_".join(part for part in normalized.split("_") if part)


def _cumpleanos_en_anio(fecha_nacimiento, anio):
    if not fecha_nacimiento:
        return None
    try:
        return fecha_nacimiento.replace(year=anio)
    except ValueError:
        return date(anio, 2, 28)


def _cumpleanos_en_rango(fecha_nacimiento, fecha_desde=None, fecha_hasta=None):
    if not fecha_nacimiento:
        return False
    md = fecha_nacimiento.month * 100 + fecha_nacimiento.day
    if fecha_desde:
        md_desde = fecha_desde.month * 100 + fecha_desde.day
        if md < md_desde:
            return False
    if fecha_hasta:
        md_hasta = fecha_hasta.month * 100 + fecha_hasta.day
        if md > md_hasta:
            return False
    return True


def _estado_vencimiento_documento(fecha_vencimiento, fecha_referencia=None, dias_alerta=30):
    if not fecha_vencimiento:
        return "Sin fecha"
    fecha_referencia = fecha_referencia or timezone.localdate()
    if fecha_vencimiento <= fecha_referencia:
        return "Vencido"
    return "Vigente"


def _apply_date_range(queryset, field_name, fecha_desde=None, fecha_hasta=None):
    filters = {}
    if fecha_desde:
        filters[f"{field_name}__gte"] = fecha_desde
    if fecha_hasta:
        filters[f"{field_name}__lte"] = fecha_hasta
    return queryset.filter(**filters) if filters else queryset


def _filter_period_overlap(queryset, start_field, end_field, fecha_desde=None, fecha_hasta=None):
    if fecha_desde:
        queryset = queryset.filter(**{f"{end_field}__gte": fecha_desde})
    if fecha_hasta:
        queryset = queryset.filter(**{f"{start_field}__lte": fecha_hasta})
    return queryset


def _export_report_to_excel(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte"

    sheet["A1"] = report["title"]
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = report["description"]
    sheet["A2"].font = Font(italic=True, size=10)

    header_row = 4
    fill = PatternFill("solid", fgColor="DCEBFF")
    for column_index, column_name in enumerate(report["columns"], start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = fill

    for row_index, row in enumerate(report["rows"], start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    footer_row = header_row + len(report["rows"]) + 2
    sheet.cell(row=footer_row, column=1, value=report["total_label"]).font = Font(bold=True)
    sheet.cell(row=footer_row, column=2, value=str(report["total_value"])).font = Font(bold=True)

    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 42)

    filename = f"{_slugify_filename(report['title'])}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _build_report(tipo_reporte, fecha_desde=None, fecha_hasta=None):
    hoy = timezone.localdate()
    Vacacion.sincronizar_estados()

    if tipo_reporte == "choferes_registrados":
        registros = list(
            _apply_date_range(Chofer.objects.all(), "fecha_registro", fecha_desde, fecha_hasta)
            .order_by("nombre")
            .values("nombre", "cedula", "licencia", "carta_buena_conducta", "rntt")
        )
        return {
            "title": "Choferes subcontratistas registrados",
            "description": "Listado general de choferes subcontratistas disponibles en el sistema.",
            "columns": ["Nombre", "Cedula", "Licencia", "Carta buena conducta", "RNTT"],
            "rows": [
                [
                    item["nombre"],
                    item["cedula"],
                    item["licencia"],
                    "Si" if item["carta_buena_conducta"] else "No",
                    "Si" if item["rntt"] else "No",
                ]
                for item in registros
            ],
            "total_label": "Choferes encontrados",
            "total_value": len(registros),
        }

    if tipo_reporte == "empleados_activos":
        registros = list(
            _apply_date_range(
                Empleado.objects.filter(estado=Empleado.Estado.ACTIVO),
                "fecha_ingreso",
                fecha_desde,
                fecha_hasta,
            )
            .select_related("cargo", "cargo__departamento")
            .order_by("nombre")
        )
        return {
            "title": "Empleados activos",
            "description": "Personal interno activo actualmente en la empresa.",
            "columns": ["Nombre", "Cedula", "Cargo", "Departamento", "Ingreso"],
            "rows": [
                [
                    f"{item.nombre} {item.apellidos}".strip(),
                    item.cedula,
                    item.cargo.nombre,
                    item.cargo.departamento.nombre,
                    item.fecha_ingreso.strftime("%d/%m/%Y"),
                ]
                for item in registros
            ],
            "total_label": "Empleados activos",
            "total_value": len(registros),
        }

    if tipo_reporte == "choferes_por_estado":
        registros = list(
            _apply_date_range(Chofer.objects.all(), "fecha_registro", fecha_desde, fecha_hasta)
            .values("estado")
            .annotate(total=Count("id"))
            .order_by("estado")
        )
        return {
            "title": "Choferes por estado",
            "description": "Resumen de choferes subcontratistas agrupados por estado operativo.",
            "columns": ["Estado", "Cantidad"],
            "rows": [[_choice_label(Chofer.Estado.choices, item["estado"]), item["total"]] for item in registros],
            "total_label": "Choferes registrados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "choferes_carta_buena_conducta":
        registros = list(
            _apply_date_range(Chofer.objects.all(), "fecha_registro", fecha_desde, fecha_hasta)
            .values("carta_buena_conducta")
            .annotate(total=Count("id"))
            .order_by("carta_buena_conducta")
        )
        return {
            "title": "Choferes por carta de buena conducta",
            "description": "Resumen de choferes segun disponibilidad de carta de buena conducta.",
            "columns": ["Carta de buena conducta", "Cantidad"],
            "rows": [["Si" if item["carta_buena_conducta"] else "No", item["total"]] for item in registros],
            "total_label": "Choferes contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "choferes_rntt":
        registros = list(
            _apply_date_range(Chofer.objects.all(), "fecha_registro", fecha_desde, fecha_hasta)
            .values("rntt")
            .annotate(total=Count("id"))
            .order_by("rntt")
        )
        return {
            "title": "Choferes por cumplimiento de RNTT",
            "description": "Resumen de choferes segun cumplimiento del requisito interno RNTT.",
            "columns": ["RNTT", "Cantidad"],
            "rows": [["Si" if item["rntt"] else "No", item["total"]] for item in registros],
            "total_label": "Choferes contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "choferes_licencias_por_vencimiento":
        choferes = _apply_date_range(
            Chofer.objects.filter(vencimiento_licencia__isnull=False),
            "vencimiento_licencia",
            fecha_desde,
            fecha_hasta,
        ).order_by("vencimiento_licencia", "nombre")
        registros = list(choferes)
        return {
            "title": "Licencias de choferes por vencimiento",
            "description": "Control de vencimiento de licencias de conducir de choferes subcontratistas.",
            "columns": ["Chofer", "Cedula", "Licencia", "Categoria", "Vencimiento", "Estado"],
            "rows": [
                [
                    item.nombre,
                    item.cedula,
                    item.licencia,
                    item.categoria_licencia or "N/D",
                    item.vencimiento_licencia.strftime("%d/%m/%Y"),
                    _estado_vencimiento_documento(item.vencimiento_licencia, hoy),
                ]
                for item in registros
            ],
            "total_label": "Choferes listados",
            "total_value": len(registros),
        }

    if tipo_reporte == "choferes_rntt_por_vencimiento":
        choferes = _apply_date_range(
            Chofer.objects.filter(rntt=True, vencimiento_carnet_rntt__isnull=False),
            "vencimiento_carnet_rntt",
            fecha_desde,
            fecha_hasta,
        ).order_by("vencimiento_carnet_rntt", "nombre")
        registros = list(choferes)
        return {
            "title": "Carnet RNTT por vencimiento",
            "description": "Seguimiento de vencimiento del carnet RNTT de choferes activos en el requisito.",
            "columns": ["Chofer", "Cedula", "RNTT", "Vencimiento carnet", "Estado"],
            "rows": [
                [
                    item.nombre,
                    item.cedula,
                    "Si" if item.rntt else "No",
                    item.vencimiento_carnet_rntt.strftime("%d/%m/%Y"),
                    _estado_vencimiento_documento(item.vencimiento_carnet_rntt, hoy),
                ]
                for item in registros
            ],
            "total_label": "Choferes listados",
            "total_value": len(registros),
        }

    if tipo_reporte == "choferes_seguro_ley_por_vencimiento":
        choferes = _apply_date_range(
            Chofer.objects.filter(seguro_ley=True, vencimiento_seguro_ley__isnull=False),
            "vencimiento_seguro_ley",
            fecha_desde,
            fecha_hasta,
        ).order_by("vencimiento_seguro_ley", "nombre")
        registros = list(choferes)
        return {
            "title": "Seguro de ley por vencimiento",
            "description": "Control de vencimiento del seguro de ley de choferes subcontratistas.",
            "columns": ["Chofer", "Cedula", "Seguro de ley", "Vencimiento seguro", "Estado"],
            "rows": [
                [
                    item.nombre,
                    item.cedula,
                    "Si" if item.seguro_ley else "No",
                    item.vencimiento_seguro_ley.strftime("%d/%m/%Y"),
                    _estado_vencimiento_documento(item.vencimiento_seguro_ley, hoy),
                ]
                for item in registros
            ],
            "total_label": "Choferes listados",
            "total_value": len(registros),
        }

    if tipo_reporte == "conduces_listado":
        registros = list(
            _apply_date_range(
                Conduce.objects.select_related("chofer", "vehiculo"),
                "fecha",
                fecha_desde,
                fecha_hasta,
            ).order_by("-fecha", "-id")
        )
        return {
            "title": "Listado de conduces",
            "description": "Listado detallado de conduces registrados en el periodo seleccionado.",
            "columns": [
                "Numero de conduce",
                "Fecha",
                "Chofer",
                "Vehiculo",
                "Origen",
                "Destino",
                "Monto generado",
                "Estado",
            ],
            "rows": [
                [
                    item.numero,
                    item.fecha.strftime("%d/%m/%Y") if item.fecha else "N/D",
                    item.chofer.nombre,
                    item.vehiculo.placa if item.vehiculo else "N/D",
                    item.origen or "N/D",
                    item.destino or "N/D",
                    _format_currency(item.monto_generado or 0),
                    item.get_estado_display(),
                ]
                for item in registros
            ],
            "total_label": "Conduces listados",
            "total_value": len(registros),
        }

    if tipo_reporte == "licencias_activas":
        registros = list(
            _filter_period_overlap(
                Licencia.objects.select_related("empleado", "tipo")
                .filter(fecha_inicio__lte=hoy, fecha_fin__gte=hoy),
                "fecha_inicio",
                "fecha_fin",
                fecha_desde,
                fecha_hasta,
            )
            .order_by("fecha_inicio")
        )
        return {
            "title": "Licencias activas",
            "description": "Colaboradores que se encuentran actualmente bajo licencia.",
            "columns": ["Empleado", "Tipo", "Inicio", "Fin", "Estado"],
            "rows": [
                [
                    f"{item.empleado.nombre} {item.empleado.apellidos}".strip(),
                    item.tipo.nombre,
                    item.fecha_inicio.strftime("%d/%m/%Y"),
                    item.fecha_fin.strftime("%d/%m/%Y"),
                    item.get_estado_display(),
                ]
                for item in registros
            ],
            "total_label": "Licencias activas",
            "total_value": len(registros),
        }

    if tipo_reporte == "empleados_por_estado":
        registros = list(
            _apply_date_range(Empleado.objects.all(), "fecha_ingreso", fecha_desde, fecha_hasta)
            .values("estado")
            .annotate(total=Count("id"))
            .order_by("estado")
        )
        return {
            "title": "Empleados por estado",
            "description": "Clasificacion del personal interno por estado laboral.",
            "columns": ["Estado", "Cantidad"],
            "rows": [[_choice_label(Empleado.Estado.choices, item["estado"]), item["total"]] for item in registros],
            "total_label": "Empleados contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "empleados_por_departamento":
        registros = list(
            _apply_date_range(Empleado.objects.all(), "fecha_ingreso", fecha_desde, fecha_hasta)
            .values("cargo__departamento__nombre")
            .annotate(total=Count("id"))
            .order_by("cargo__departamento__nombre")
        )
        return {
            "title": "Empleados por departamento",
            "description": "Distribucion del personal interno por departamento.",
            "columns": ["Departamento", "Cantidad"],
            "rows": [[item["cargo__departamento__nombre"] or "Sin departamento", item["total"]] for item in registros],
            "total_label": "Empleados contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "empleados_por_cumpleanos":
        empleados = list(
            Empleado.objects.select_related("cargo", "cargo__departamento")
            .filter(fecha_nacimiento__isnull=False)
            .order_by("nombre", "apellidos")
        )
        anio_referencia = (fecha_desde or fecha_hasta or hoy).year
        registros = []
        for empleado in empleados:
            if not _cumpleanos_en_rango(empleado.fecha_nacimiento, fecha_desde, fecha_hasta):
                continue
            cumple_anual = _cumpleanos_en_anio(empleado.fecha_nacimiento, anio_referencia)
            if not cumple_anual:
                continue
            registros.append(
                {
                    "empleado": f"{empleado.nombre} {empleado.apellidos}".strip(),
                    "cedula": empleado.cedula,
                    "departamento": empleado.cargo.departamento.nombre,
                    "fecha_nacimiento": empleado.fecha_nacimiento,
                    "cumple_anual": cumple_anual,
                }
            )

        registros.sort(key=lambda item: item["cumple_anual"])
        return {
            "title": "Empleados por fecha de cumpleanos",
            "description": "Listado de empleados filtrado por rango de cumpleanos (mes y dia).",
            "columns": ["Empleado", "Cedula", "Departamento", "Fecha de nacimiento", "Cumpleanos"],
            "rows": [
                [
                    item["empleado"],
                    item["cedula"],
                    item["departamento"],
                    item["fecha_nacimiento"].strftime("%d/%m/%Y"),
                    item["cumple_anual"].strftime("%d/%m"),
                ]
                for item in registros
            ],
            "total_label": "Empleados listados",
            "total_value": len(registros),
        }

    if tipo_reporte == "gastos_rrhh_listado":
        registros = list(
            _apply_date_range(RegistroGasto.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .order_by("-fecha", "-id")
        )
        return {
            "title": "Todos los gastos (listado)",
            "description": "Listado detallado de gastos de RRHH en el rango de fechas seleccionado.",
            "columns": [
                "Fecha",
                "Proveedor",
                "Comprobante",
                "Numero comprobante",
                "Valor",
                "ITBIS",
                "Propinas",
                "Total",
                "Motivo",
            ],
            "rows": [
                [
                    item.fecha.strftime("%d/%m/%Y"),
                    item.proveedor,
                    "Con comprobante" if item.con_comprobante else "Sin comprobante",
                    item.numero_comprobante or "-",
                    _format_currency(item.valor or 0),
                    _format_currency(item.itbis or 0),
                    _format_currency(item.propinas or 0),
                    _format_currency(item.total),
                    item.motivo,
                ]
                for item in registros
            ],
            "total_label": "Total gastado",
            "total_value": _format_currency(sum((item.total or 0) for item in registros)),
        }

    if tipo_reporte == "gastos_rrhh_por_proveedor":
        registros = list(
            _apply_date_range(RegistroGasto.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .values("proveedor")
            .annotate(
                cantidad=Count("id"),
                valor=Sum("valor"),
                itbis=Sum("itbis"),
                propinas=Sum("propinas"),
            )
            .order_by("proveedor")
        )
        return {
            "title": "Gastos por proveedor",
            "description": "Consolidado de gastos de RRHH agrupado por proveedor.",
            "columns": ["Proveedor", "Registros", "Valor", "ITBIS", "Propinas", "Total"],
            "rows": [
                [
                    item["proveedor"],
                    item["cantidad"],
                    _format_currency(item["valor"] or 0),
                    _format_currency(item["itbis"] or 0),
                    _format_currency(item["propinas"] or 0),
                    _format_currency(
                        (item["valor"] or 0) + (item["itbis"] or 0) + (item["propinas"] or 0)
                    ),
                ]
                for item in registros
            ],
            "total_label": "Total gastado",
            "total_value": _format_currency(
                sum(
                    (item["valor"] or 0) + (item["itbis"] or 0) + (item["propinas"] or 0)
                    for item in registros
                )
            ),
        }

    if tipo_reporte == "licencias_por_estado":
        registros = list(
            _filter_period_overlap(Licencia.objects.all(), "fecha_inicio", "fecha_fin", fecha_desde, fecha_hasta)
            .values("estado")
            .annotate(total=Count("id"))
            .order_by("estado")
        )
        return {
            "title": "Licencias por estado",
            "description": "Resumen de licencias registradas segun su estado.",
            "columns": ["Estado", "Cantidad"],
            "rows": [[_choice_label(Licencia.Estado.choices, item["estado"]), item["total"]] for item in registros],
            "total_label": "Licencias contabilizadas",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "tipos_licencia_configurados":
        registros = list(
            TipoLicencia.objects.values("nombre", "requiere_aprobacion", "activo")
            .annotate(total=Count("id"))
            .order_by("nombre")
        )
        return {
            "title": "Tipos de licencia configurados",
            "description": "Catalogo de tipos de licencia configurados en RRHH.",
            "columns": ["Tipo de licencia", "Requiere aprobacion", "Activo"],
            "rows": [
                [item["nombre"], "Si" if item["requiere_aprobacion"] else "No", "Si" if item["activo"] else "No"]
                for item in registros
            ],
            "total_label": "Tipos configurados",
            "total_value": len(registros),
        }

    if tipo_reporte == "vacaciones_activas":
        registros = list(
            _filter_period_overlap(
                Vacacion.objects.select_related("empleado")
                .filter(fecha_inicio__lte=hoy, fecha_fin__gte=hoy),
                "fecha_inicio",
                "fecha_fin",
                fecha_desde,
                fecha_hasta,
            )
            .order_by("fecha_inicio")
        )
        return {
            "title": "Vacaciones activas",
            "description": "Empleados que actualmente se encuentran de vacaciones.",
            "columns": ["Empleado", "Inicio", "Fin", "Estado"],
            "rows": [
                [
                    f"{item.empleado.nombre} {item.empleado.apellidos}".strip(),
                    item.fecha_inicio.strftime("%d/%m/%Y"),
                    item.fecha_fin.strftime("%d/%m/%Y"),
                    item.get_estado_display(),
                ]
                for item in registros
            ],
            "total_label": "Vacaciones activas",
            "total_value": len(registros),
        }

    if tipo_reporte == "vacaciones_por_estado":
        registros = list(
            _filter_period_overlap(Vacacion.objects.all(), "fecha_inicio", "fecha_fin", fecha_desde, fecha_hasta)
            .values("estado")
            .annotate(total=Count("id"))
            .order_by("estado")
        )
        return {
            "title": "Vacaciones por estado",
            "description": "Resumen de vacaciones segun su estado de ejecucion.",
            "columns": ["Estado", "Cantidad"],
            "rows": [[_choice_label(Vacacion.Estado.choices, item["estado"]), item["total"]] for item in registros],
            "total_label": "Vacaciones contabilizadas",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "pagos_empleados_por_mes":
        registros = list(
            _apply_date_range(PagoEmpleado.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .annotate(periodo=TruncMonth("fecha"))
            .values("periodo")
            .annotate(total=Sum("monto"), cantidad=Count("id"))
            .order_by("-periodo")
        )
        return {
            "title": "Pagos de empleados por mes",
            "description": "Consolidado mensual de pagos realizados al personal interno.",
            "columns": ["Mes", "Total pagado"],
            "rows": [
                [
                    item["periodo"].strftime("%m/%Y") if item["periodo"] else "Sin fecha",
                    _format_currency(item["total"] or 0),
                ]
                for item in registros
            ],
            "total_label": "Total general pagado",
            "total_value": _format_currency(
                _apply_date_range(PagoEmpleado.objects.all(), "fecha", fecha_desde, fecha_hasta).aggregate(total=Sum("monto"))["total"] or 0
            ),
        }

    if tipo_reporte == "pagos_empleados_por_metodo":
        registros = list(
            _apply_date_range(PagoEmpleado.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .values("metodo")
            .annotate(total_pagado=Sum("monto"), cantidad=Count("id"))
            .order_by("metodo")
        )
        return {
            "title": "Pagos de empleados por metodo",
            "description": "Resumen de pagos de empleados por metodo de desembolso.",
            "columns": ["Metodo", "Cantidad", "Total pagado"],
            "rows": [
                [_choice_label(PagoEmpleado.Metodo.choices, item["metodo"]), item["cantidad"], _format_currency(item["total_pagado"] or 0)]
                for item in registros
            ],
            "total_label": "Total general pagado",
            "total_value": _format_currency(sum((item["total_pagado"] or 0) for item in registros)),
        }

    if tipo_reporte == "pagos_por_mes":
        registros = list(
            _apply_date_range(Pago.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .annotate(periodo=TruncMonth("fecha"))
            .values("periodo")
            .annotate(total=Sum("monto"), cantidad=Count("id"))
            .order_by("-periodo")
        )
        return {
            "title": "Total pagado a choferes por mes",
            "description": "Consolidado mensual de pagos realizados a choferes subcontratistas.",
            "columns": ["Mes", "Total pagado"],
            "rows": [
                [
                    item["periodo"].strftime("%m/%Y") if item["periodo"] else "Sin fecha",
                    _format_currency(item["total"] or 0),
                ]
                for item in registros
            ],
            "total_label": "Total general pagado",
            "total_value": _format_currency(
                _apply_date_range(Pago.objects.all(), "fecha", fecha_desde, fecha_hasta).aggregate(total=Sum("monto"))["total"] or 0
            ),
        }

    if tipo_reporte == "pagos_por_metodo":
        registros = list(
            _apply_date_range(Pago.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .values("metodo")
            .annotate(total_pagado=Sum("monto"), cantidad=Count("id"))
            .order_by("metodo")
        )
        return {
            "title": "Pagos a choferes por metodo",
            "description": "Resumen de pagos agrupados por metodo de desembolso.",
            "columns": ["Metodo", "Cantidad", "Total pagado"],
            "rows": [[_choice_label(Pago.Metodo.choices, item["metodo"]), item["cantidad"], _format_currency(item["total_pagado"] or 0)] for item in registros],
            "total_label": "Total general pagado",
            "total_value": _format_currency(sum((item["total_pagado"] or 0) for item in registros)),
        }

    if tipo_reporte == "productos_por_categoria":
        registros = list(
            Producto.objects.values("categoria")
            .annotate(total=Count("id"), stock=Sum("cantidad"))
            .order_by("categoria")
        )
        return {
            "title": "Productos por categoria",
            "description": "Clasificacion de inventario segun categoria de producto.",
            "columns": ["Categoria", "Productos", "Stock total"],
            "rows": [[_choice_label(Producto.Categoria.choices, item["categoria"]), item["total"], item["stock"] or 0] for item in registros],
            "total_label": "Productos registrados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "productos_activos":
        activos = Producto.objects.filter(activo=True).count()
        inactivos = Producto.objects.filter(activo=False).count()
        return {
            "title": "Productos activos e inactivos",
            "description": "Resumen del inventario segun disponibilidad operativa del producto.",
            "columns": ["Estado", "Cantidad"],
            "rows": [["Activos", activos], ["Inactivos", inactivos]],
            "total_label": "Productos contabilizados",
            "total_value": activos + inactivos,
        }

    if tipo_reporte == "movimientos_por_tipo":
        registros = list(
            _apply_date_range(MovimientoInventario.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .values("tipo")
            .annotate(total=Count("id"), cantidad_total=Sum("cantidad"))
            .order_by("tipo")
        )
        return {
            "title": "Movimientos de inventario por tipo",
            "description": "Resumen de entradas, salidas y ajustes del inventario.",
            "columns": ["Tipo", "Movimientos", "Cantidad total"],
            "rows": [[_choice_label(MovimientoInventario.Tipo.choices, item["tipo"]), item["total"], item["cantidad_total"] or 0] for item in registros],
            "total_label": "Movimientos contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "suministros_combustible_por_estado":
        registros = list(
            _apply_date_range(AvanceChofer.objects.all(), "fecha", fecha_desde, fecha_hasta)
            .values("estado")
            .annotate(
                total=Count("id"),
                galones=Sum("galones"),
                monto=Sum("monto"),
                saldo=Sum("saldo_pendiente"),
            )
            .order_by("estado")
        )
        return {
            "title": "Suministros de combustible por estado",
            "description": "Resumen de suministros de combustible entregados a choferes.",
            "columns": ["Estado", "Suministros", "Galones", "Monto suministrado", "Saldo pendiente"],
            "rows": [
                [
                    _choice_label(AvanceChofer.Estado.choices, item["estado"]),
                    item["total"],
                    float(item["galones"] or 0),
                    _format_currency(item["monto"] or 0),
                    _format_currency(item["saldo"] or 0),
                ]
                for item in registros
            ],
            "total_label": "Monto total suministrado",
            "total_value": _format_currency(sum((item["monto"] or 0) for item in registros)),
        }

    if tipo_reporte == "vehiculos_disponibles":
        registros = list(
            Vehiculo.objects.filter(estado=Vehiculo.Estado.DISPONIBLE)
            .order_by("placa")
            .values(
                "placa",
                "marca",
                "modelo",
                "ultima_ubicacion",
                "es_propiedad_empresa",
                "dueno_nombre",
            )
        )
        return {
            "title": "Vehiculos disponibles",
            "description": "Unidades actualmente disponibles para asignacion o servicio.",
            "columns": ["Placa", "Marca", "Modelo", "Ultima ubicacion", "Propiedad / dueno"],
            "rows": [
                [
                    item["placa"],
                    item["marca"] or "N/D",
                    item["modelo"],
                    item["ultima_ubicacion"] or "N/D",
                    "Empresa" if item["es_propiedad_empresa"] else f"Alquilado - {item['dueno_nombre'] or 'Dueno no especificado'}",
                ]
                for item in registros
            ],
            "total_label": "Vehiculos disponibles",
            "total_value": len(registros),
        }

    if tipo_reporte == "vehiculos_por_estado":
        registros = list(Vehiculo.objects.values("estado").annotate(total=Count("id")).order_by("estado"))
        return {
            "title": "Vehiculos por estado",
            "description": "Resumen de la flota segun el estado operativo de cada vehiculo.",
            "columns": ["Estado", "Cantidad"],
            "rows": [[_choice_label(Vehiculo.Estado.choices, item["estado"]), item["total"]] for item in registros],
            "total_label": "Vehiculos contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "contenedores_disponibles":
        registros = list(
            Contenedor.objects.filter(estado=Contenedor.Estado.DISPONIBLE)
            .order_by("codigo")
            .values("codigo", "color")
        )
        return {
            "title": "Contenedores disponibles",
            "description": "Listado de contenedores disponibles para nuevas asignaciones.",
            "columns": ["Ficha del contenedor", "Color", "Estado"],
            "rows": [[item["codigo"], item["color"] or "N/D", "Disponible"] for item in registros],
            "total_label": "Contenedores disponibles",
            "total_value": len(registros),
        }

    if tipo_reporte == "contenedores_por_estado":
        registros = list(
            Contenedor.objects.values("estado").annotate(total=Count("id")).order_by("estado")
        )
        return {
            "title": "Contenedores por estado",
            "description": "Resumen de contenedores segun estado de disponibilidad.",
            "columns": ["Estado", "Cantidad"],
            "rows": [[_choice_label(Contenedor.Estado.choices, item["estado"]), item["total"]] for item in registros],
            "total_label": "Contenedores contabilizados",
            "total_value": sum(item["total"] for item in registros),
        }

    if tipo_reporte == "contenedores_alquilados":
        registros = list(
            _apply_date_range(
                AlquilerContenedor.objects.select_related("contenedor", "chasis")
                .filter(estado=AlquilerContenedor.Estado.ACTIVO),
                "fecha_inicio",
                fecha_desde,
                fecha_hasta,
            )
            .order_by("fecha_inicio", "id")
            .values(
                "contenedor__codigo",
                "contenedor__color",
                "cliente",
                "fecha_inicio",
                "fecha_fin",
                "con_chasis",
                "chasis__codigo",
            )
        )
        return {
            "title": "Contenedores alquilados",
            "description": "Contenedores actualmente alquilados y su periodo de alquiler.",
            "columns": [
                "Ficha del contenedor",
                "Color",
                "Alquilado a",
                "Fecha inicio",
                "Fecha fin",
                "Con chasis",
                "Chasis",
            ],
            "rows": [
                [
                    item["contenedor__codigo"],
                    item["contenedor__color"] or "N/D",
                    item["cliente"] or "No definido",
                    item["fecha_inicio"].strftime("%d/%m/%Y") if item["fecha_inicio"] else "N/D",
                    item["fecha_fin"].strftime("%d/%m/%Y") if item["fecha_fin"] else "N/D",
                    "Si" if item["con_chasis"] else "No",
                    item["chasis__codigo"] or "N/D",
                ]
                for item in registros
            ],
            "total_label": "Contenedores alquilados",
            "total_value": len(registros),
        }

    if tipo_reporte == "registros_proveedores_por_empresa":
        registros = list(
            _apply_date_range(RegistroProveedor.objects.select_related("proveedor"), "fecha", fecha_desde, fecha_hasta)
            .values("proveedor__nombre")
            .annotate(total_registros=Count("id"), total_tarifa=Sum("tarifa"))
            .order_by("proveedor__nombre")
        )
        return {
            "title": "Registros de proveedores por empresa",
            "description": "Consolidado por empresa proveedora con cantidad de registros y tarifa acumulada.",
            "columns": ["Empresa", "Registros", "Tarifa acumulada"],
            "rows": [
                [
                    item["proveedor__nombre"],
                    item["total_registros"],
                    _format_currency(item["total_tarifa"] or 0),
                ]
                for item in registros
            ],
            "total_label": "Tarifa total consolidada",
            "total_value": _format_currency(sum((item["total_tarifa"] or 0) for item in registros)),
        }

    if tipo_reporte == "saldo_combustible_pendiente":
        registros = list(
            _apply_date_range(
                AvanceChofer.objects.select_related("chofer")
                .filter(saldo_pendiente__gt=0),
                "fecha",
                fecha_desde,
                fecha_hasta,
            )
            .order_by("-fecha", "-id")
        )
        return {
            "title": "Saldo pendiente por combustible suministrado por adelantado",
            "description": "Suministros de combustible por adelantado pendientes de cobro a choferes.",
            "columns": ["Fecha", "Chofer", "Galones", "Precio por galon", "Monto suministro", "Saldo pendiente"],
            "rows": [
                [
                    item.fecha.strftime("%d/%m/%Y"),
                    item.chofer.nombre,
                    float(item.galones or 0),
                    _format_currency(item.precio_por_galon or 0),
                    _format_currency(item.monto or 0),
                    _format_currency(item.saldo_pendiente),
                ]
                for item in registros
            ],
            "total_label": "Saldo pendiente total",
            "total_value": _format_currency(
                _apply_date_range(
                    AvanceChofer.objects.filter(saldo_pendiente__gt=0),
                    "fecha",
                    fecha_desde,
                    fecha_hasta,
                ).aggregate(total=Sum("saldo_pendiente"))["total"] or 0
            ),
        }

    if tipo_reporte == "cobros_combustible_por_mes":
        registros = list(
            _apply_date_range(Pago.objects.filter(descuento_avances__gt=0), "fecha", fecha_desde, fecha_hasta)
            .annotate(periodo=TruncMonth("fecha"))
            .values("periodo")
            .annotate(total=Sum("descuento_avances"), cantidad=Count("id"))
            .order_by("-periodo")
        )
        return {
            "title": "Cobros de combustible por mes",
            "description": "Cobros aplicados por suministro de combustible en pagos a choferes.",
            "columns": ["Mes", "Pagos con descuento", "Total cobrado"],
            "rows": [
                [
                    item["periodo"].strftime("%m/%Y") if item["periodo"] else "Sin fecha",
                    item["cantidad"],
                    _format_currency(item["total"] or 0),
                ]
                for item in registros
            ],
            "total_label": "Total cobrado por combustible",
            "total_value": _format_currency(sum((item["total"] or 0) for item in registros)),
        }

    return None


def lista_reportes(request):
    form = GeneradorReporteForm(request.GET or None)
    selected_report = None

    if form.is_valid():
        selected_report = _build_report(
            form.cleaned_data["tipo_reporte"],
            form.cleaned_data.get("fecha_desde"),
            form.cleaned_data.get("fecha_hasta"),
        )
        if request.GET.get("action") == "excel" and selected_report:
            return _export_report_to_excel(selected_report)

    return render(
        request,
        "reportes/lista_reportes.html",
        {
            "page_title": "Reportes",
            "page_intro": "Genera reportes operativos, financieros y de personal desde un solo panel.",
            "summary_cards": [
                {"label": "Choferes", "value": Chofer.objects.count(), "accent": "blue"},
                {"label": "Pagos choferes", "value": Pago.objects.count(), "accent": "green"},
                {"label": "Pagos empleados", "value": PagoEmpleado.objects.count(), "accent": "teal"},
                {"label": "Gastos RRHH", "value": RegistroGasto.objects.count(), "accent": "amber"},
                {"label": "Vehiculos", "value": Vehiculo.objects.count(), "accent": "amber"},
                {"label": "Contenedores", "value": Contenedor.objects.count(), "accent": "blue"},
            ],
            "form": form,
            "selected_report": selected_report,
        },
    )
